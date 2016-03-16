import yaml, openpyxl, re, argparse, sys
# re for the regular expression input testing 'Match'
# argpars todo: add the input xlsx as an argument
# openpyxl enables the WorkBook manipulation
# import yaml - use this for exporting yaml
# sys enables the sys.exit()


from openpyxl import workbook,load_workbook
filename = 'rule.xlsx'
wb = load_workbook(filename, use_iterators=True)
ws = wb.get_sheet_by_name('rule')

#Get the temp files setup
head_out_file = open('build_head.yaml', "w")
ingress_out_file = open('ingress_rules.yaml', "w")
egress_out_file = open('egress_rules.yaml', "w")
head_out_file.truncate() # clear the file contents
ingress_out_file.truncate() # clear the file contents
egress_out_file.truncate()  # clear the file content
ingress_out_file.write('ingress_rules:\n') # set Yaml Variable Hash name in the file
egress_out_file.write('egress_rules:\n')   # set Yaml Variable Hash name in the file 

try:
    #Take in the boilerplate of Region, VPC ID, 
    super_region = ws['B2'].value
    #Validate that Region is set to us-east-1 
    match = re.match('^us-east-1', super_region)
    if match:
        print ('You are using Region = %s' % super_region)
    else:
        print ('You did NOT list us-east-1, but %s instead' % super_region)
        sys.exit()
    
    super_vpc_id = ws['C2'].value
    #Validate that vpc input starts with "vpc_" and has 9digits or chars after it
    match = re.match("^vpc-[\\'a-zA-Z0-9\\']{8}", super_vpc_id)
    if match:
        print ('You are using VPC-ID = %s' % super_vpc_id)
    else:
        print ('You did NOT a valid VPC-ID, %s' % super_vpc_id)
        sys.exit()
    
    super_name = ws['E2'].value
    super_description = ws['F2'].value
    
    #Write the BoilerPlate for Ansible
    head_out_file.write("---\n- hosts: localhost\n  tasks:\n    - include_vars: ingress_rules.yaml\n    - include_vars: egress_rules.yaml\n    - name: Create Groups\n      ec2_group:\n")
    head_out_file.write("        region: %s\n" % (super_region))
    head_out_file.write("        vpc_id: %s\n" % (super_vpc_id))
    head_out_file.write("        name: %s\n" % (super_name))
    head_out_file.write("        description: %s\n" % (super_description))
    head_out_file.write("        rules: \"{{ ingress_rules }}\"\n")
    head_out_file.write("        rules_egress: \"{{ egress_rules }}\"\n")
    
    for row in range(2, ws.max_row + 1): #Skip the first row of headers
        #Parse the cell input and convert to strings this notation is provided by the openpyxl library
#        current_row = row
        print ("\nWorking with Row # %s" % (row))
        super_direction = ws['G' + str(row)].value # take in the value from the spreadsheet
        match = re.match('egress|ingress', super_direction) #set the test flag
        if match:  #loop to tell the user
            print ('Direction is good')
        else:
            print ("Direction is NOT good.  You gave... %s. Must be ingress or egress" % (super_direction))
            break #ABORT if it fails
        super_proto = ws['H' + str(row)].value
        match = re.match('udp|tcp', super_proto)
        if match:
            print ('Protocol is good')
        else: 
            print ("Protocol is NOT good.  You gave... %s.  Must Be tcp or udp" % (super_proto))
            break
        #take in and and test the From Port
        super_from_port = ws['I' + str(row)].value
#        print (super_from_port)
        super_from_port = str(super_from_port)
        match = re.match('[0-9]+', super_from_port)
        if match:
            print ("From Port is good")
        else:
            print ("From Port is NOT good.  You gave %s.  It needs to be a number" % (super_from_port))
            break
        super_to_port = ws['J' + str(row)].value
        super_to_port = str(super_to_port) # digits have to be set to string for the re.match to work
        match = re.match('[0-9]+', super_to_port)
        if match:
            print ("To Port is good")
        else:
            print ("To Port is NOT good.  You gave %s.  It needs to be a number" % (super_to_port))
            break
        super_cidr_ip = ws['K' + str(row)].value
        match = re.match('^(([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])\.){3}([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])(\/([0-9]|[1-2][0-9]|3[0-2]))$', super_cidr_ip )
        if match:
            print ("Cidr IP is good")
        else:
            print ("CidrIP is NOT good %s" % (super_cidr_ip))
            break
        
        if super_direction == "ingress":
            #this line includes secret sauce of whitespace and substitution to build the output 
            ingress_out_file.write("      - proto: %s\n        from_port: %s\n        to_port: %s\n        cidr_ip: \"%s\" \n" % (super_proto, super_from_port, super_to_port, super_cidr_ip))
        elif super_direction == "egress":
            egress_out_file.write("      - proto: %s\n        from_port: %s\n        to_port: %s\n        cidr_ip: \"%s\" \n" % (super_proto, super_from_port, super_to_port, super_cidr_ip))
        else:
            print ("Check the direction, must be ingress or egress")
        

finally:
    head_out_file.close()
    ingress_out_file.close()
    egress_out_file.close()
