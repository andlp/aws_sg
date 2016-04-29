#!/usr/bin/python

import sys, getopt, openpyxl, re, os
from openpyxl import workbook,load_workbook
from subprocess import call # used to call the build_head.sh in the OS

def main(argv):
   inputfile = ''
   inputtab = ''
   try:
      opts, args = getopt.getopt(argv,"hi:t:",["ifile=","tab="])
   except getopt.GetoptError:
      print 'AWSexelParser.py -i <inputfile> -o <Excel Tab name>'
      sys.exit(2)
   for opt, arg in opts:
      if opt == '-h':
         print 'Usage: AWSexelParser.py -i <inputfile> -t <Excel Tab name>'
         print 'Requires AWS CLi to be installed and configured with AWS keys and Region Setting'
         print
         print '******* Here is the command to test your local AWS CLI.  If this doesnt work, fix your AWS CLI *******'
         print '******* aws ec2 describe-security-groups *******'
         print
         print 'This script will takes an input xlsx file, and a Tab name from inside that xlsx file'
         print 'It will create Security group rules, ingress and egress'
         print 'This script expects the VPC and Security Group ID and Security Group Name to Pre-exist'
         sys.exit()
      elif opt in ("-i", "--ifile"):
         inputfile = arg
      elif opt in ("-t", "--tab"):
         inputtab = arg
      else:
         print 'Usage: AWSexelParser.py -i <inputfile> -t <Excel Tab name>'
         print 'Requires AWS CLi to be installed and configured with AWS keys and Region Setting'
         print
         print '******* Here is the command to test your local AWS CLI.  If this doesnt work, fix your AWS CLI *******'
         print '******* aws ec2 describe-security-groups *******'
         print
         print 'This script will takes an input xlsx file, and a Tab name from inside that xlsx file'
         print 'It will create Security group rules, ingress and egress'
         print 'This script expects the VPC and Security Group ID and Security Group Name to Pre-exist'
         break
   print 'Input file is "', inputfile
   print 'Excel Tab name is "', inputtab
   
   
   
   wb = load_workbook(inputfile, use_iterators=True)
#   ws = wb.get_sheet_by_name('Sheet1') #ToDo: Logic to set this to arbitrary WS name ^vpc-*
   ws = wb.get_sheet_by_name(inputtab)
   #Get the temp files setup
   head_out_file = open('build_head.sh', "w")
   head_out_file.truncate() # clear the file contents
   super_region = 'us-east-1'
   super_vpc_id = 'vpc-e940io34'
   for row in range(2, ws.max_row + 1): #Skip the first row of headers
       print ("\nWorking with Row # %s" % (row))
       command_string = 'aws ec2 '
       
       # Set the Direction
       super_direction = ws['F' + str(row)].value # take in the value from the spreadsheet
       match = re.match('Ingress|Egress', super_direction) #set the test flag
       if match:  #loop to tell the user
           print ('Direction is good: %s' % (super_direction))
           if super_direction == 'Ingress':
               add_opt = ('authorize-security-group-ingress ') 
               command_string += add_opt
               print ('command string is %s' % (command_string))
           elif super_direction == 'Egress':
               add_opt = ('authorize-security-group-egress ') 
               command_string += add_opt
               print ('command string is %s' % (command_string))
       else:
           print ("Direction is NOT good: %s. Must be Ingress or Egress" % (super_direction))
           break #ABORT if it fails

       #Set the Security Group Name
       # aws ec2 describe-security-groups | grep sg-4d3f1d35 
       super_sg_name = ws['A' + str(row)].value # take in the value from the spreadsheet
       match = re.match('[^\>]*', super_sg_name) #set the test flag anything that s not a < 
       if match:  #loop to tell the user
           print ('SG Name is good: %s' % (super_sg_name))
#           add_opt = ('--group-name %s ' %(super_sg_name))
#           command_string += add_opt
#           print ('command string is %s' % (command_string))
       else:
          print ("SG Name is NOT good: %s. Must be some letters or numbers" % (super_sg_name))
          break #ABORT if it fails 
       
        # Set the Security Group ID value
       super_sg_id = ws['B' + str(row)].value # take in the value from the spreadsheet
       match = re.match("^sg-[\\'a-zA-Z0-9\\']{8}", super_sg_id) #set the test flag
       if match:  #loop to tell the user
           print ('SG ID is good: %s' % (super_sg_id))
           add_opt = ('--group-id %s ' %(super_sg_id))
           command_string += add_opt
           print ('command string is %s' % (command_string))
       else:
           print ("SG ID is NOT good: %s. Must be some letters or numbers" % (super_sg_id))
           break #ABORT if it fails
           
       # Set the VPC ID value
       super_vpc_id = ws['C' + str(row)].value # take in the value from the spreadsheet
       match = re.match("^vpc-[\\'a-zA-Z0-9\\']{8}", super_vpc_id) #set the test flag
       if match:  #loop to tell the user
           print ('VPC ID is good: %s' % (super_vpc_id))
       else:
           print ("VPC ID is NOT good: %s. Must be some letters or numbers" % (super_vpc_id))
           break #ABORT if it fails
       
       # Set the Description value    
       super_description = ws['D' + str(row)].value # take in the value from the spreadsheet
       match = re.match('[^\>]*', super_description) # any Char that isnt ">"
       if match:
           print ('Description is good: %s' % (super_description))
       else: 
           print ("Description is NOT good: %s.  Must be some symbol" % (super_description))
           break    
       

       
       # Set the Protocol
       super_proto = ws['G' + str(row)].value
       match = re.match('udp|tcp', super_proto)
       if match:
            print ('Protocol is good: %s' % (super_proto))
            add_opt = ('--protocol %s ' %(super_proto)) 
            command_string += add_opt
            print ('command string is %s' % (command_string))
       else: 
           print ("Protocol is NOT good: %s.  Must be tcp or udp" % (super_proto))
           break    
                
    # Set the From_Port 
       super_from_port = ws['I' + str(row)].value
       super_from_port = str(super_from_port)
       match = re.match('[0-9]+', super_from_port)
       if match:
           print ("From Port is good: %s" % (super_from_port))
       else:
           print ("From Port is NOT good: %s.  Must be a number" % (super_from_port))
           break
       
       #Set the To_Port
       super_to_port = ws['H' + str(row)].value
       super_to_port = str(super_to_port) # digits have to be set to string for the re.match to work
       match = re.match('[0-9]+', super_to_port)
       if match:
           print ("To Port is good: %s" % (super_to_port))

       else:
           print ("To Port is NOT good: %s.  Must be a number" % (super_to_port))
           break
       
       #Set Port Output
       add_opt = ('--port %s-%s ' % (super_from_port, super_to_port)) 
       command_string += add_opt
       print ('command string is %s' % (command_string))
       
              
       # Set the Cidr IP 
       super_cidr_ip = ws['J' + str(row)].value
       super_cidr_ip = str(super_cidr_ip)
       cidr_match = re.match('^(([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])\.){3}([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])(\/([0-9]|[1-2][0-9]|3[0-2]))$', super_cidr_ip )
       print ("cider IP: %s" % (super_cidr_ip))
       if super_cidr_ip == 'None':
           print ('There is no Cidr IP in the Sheet. Checking for SG Group')
       elif cidr_match:
           print ("Cidr IP is good: %s" % (super_cidr_ip))
        
       #set the Destination Group     
       super_dest_group_id = ws['K' + str(row)].value
       super_dest_group_id = str(super_dest_group_id)
       dest_match = re.search('.* (sg-\w+).*', super_dest_group_id)
       if dest_match:
           super_dest_group_id = dest_match.group(1)
           print ("Destination Group is good: %s" % (super_dest_group_id))
       
       #Set Destination Output
       if cidr_match:
           add_opt = ('--cidr %s' % (super_cidr_ip))
           command_string += add_opt
       elif dest_match:
           add_opt = ('--source-group %s' % (super_dest_group_id))
           command_string += add_opt
             
       #Write the File Output
       head_out_file.write("\n echo AWS Appling Row %s; \n" % (row))
       head_out_file.write("%s;\n" % (command_string))

   head_out_file.close()
   print ("\n\nData Processing has completed.  Ready for the next step...")
   
   #Accept input, cast to string,
   print ("Now you have some choices...")
   print ("YES: Push the changes to AWS and Exit")
   print ("SHOW: Output the AWS commands that have been generated and Exit")
   print ("NO: Just Exit")
   do_it = raw_input('What is your choice? YES, NO, or SHOW \n')
   do_it = str(do_it)

   #Execute to AWS
   if (re.match('^yes', do_it, re.IGNORECASE)):
      print ('Proceeding to make changes in AWS.  Look here for Errors. No errors = success')
      os.system("./build_head.sh")
   #Show the AWS command file
   elif (re.match('^show', do_it, re.IGNORECASE)):
       os.system("cat ./build_head.sh")
    #Dont Execute to AWS
   else:
         print("Exiting now without changing AWS.")
         sys.exit()

if __name__ == "__main__":
   main(sys.argv[1:])
